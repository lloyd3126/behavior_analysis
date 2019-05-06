# !/usr/bin/python3
# coding:utf-8

from openpyxl import Workbook
import os
import re

from os import walk
from os.path import join

import numpy as np

# txt檔操作
# ---------start-------------

# 指定要列出所有檔案的目錄
mypath = "./output"
pathArr = []
datasheetArr = []

# 遞迴列出所有檔案的絕對路徑
for root, dirs, files in walk(mypath):
    for f in files:
        fullpath = join(root, f)
        pathArr.append(fullpath)

        fullpath = os.path.basename(fullpath).rstrip('.txt')
        fullpath = os.path.basename(fullpath).lstrip('./output/')
        datasheetArr.append(fullpath)

print(datasheetArr)

# ---------done-------------


# 定義整理資料的函式
# ---------start-------------

def clean(path):
    global columnNum
    global rowNum

    text_file = open(path, "r")
    lines = text_file.readlines()
    columnNum = len(lines)

    i = 0
    while i < columnNum:
        lines[i] = lines[i].rstrip('\n')  # 清除'\n'
        c = lines[i].split("#", 1)  # 以"#”切開分為兩個字串放入同一個陣列
        c[0] = c[0].rstrip(' ')  # 清除序列最右邊的空格
        c[0] = c[0].split(" -1",)  # 以"-1”切開分為兩個字串放入同一個陣列
        del c[0][len(c[0])-1]  # 清除陣列中最右邊的值
        j = 0
        while j < len(c[0]):
            # 正規式：清除()和[]中的數值包括符號本身
            c[0][j] = re.sub("[\(\[].*?[\)\]]", "", c[0][j])
            c[0][j] = c[0][j].replace('  ', ' ')  # 替換序列中連續的空格為單一空格
            c[0][j] = c[0][j].rstrip(' ')  # 清除序列最右邊的空格
            c[0][j] = c[0][j].lstrip(' ')  # 清除將每個序列陣列中的左側的空白消除
            j += 1
        if len(c[0]) > rowNum:  # 比較出要放在 EXCEL 哪個位置
            rowNum = len(c[0])

        clean_behaviorArr.append(c)
        i += 1
    # print(clean_behaviorArr)
    text_file.close()

# ---------done-------------


# 整理後的資料放入excel
# ---------start-------------

wb = Workbook()
dest_filename = 'output.xlsx'

for fileNum in range(0, len(datasheetArr)):
    print('Processing -------> ' + datasheetArr[fileNum] + '.txt')

    rowNum = 0
    columnNum = 0
    clean_behaviorArr = []

    # 呼叫整理資料的函式

    clean(pathArr[fileNum])
    datasheet = wb.create_sheet(datasheetArr[fileNum])

    for i in range(0, len(clean_behaviorArr)):
        for j in range(0, len(clean_behaviorArr[i][0])):
            datasheet.cell(row=i+1, column=j +
                           1).value = clean_behaviorArr[i][0][j]  # 行為
        datasheet.cell(row=i+1, column=rowNum +
                       1).value = clean_behaviorArr[i][1]  # Sup

    # ---------done-------------

    # 處理儲存格自動寬度
    # ---------start-------------

    cellinexcel = []

    # 將每個在該儲存表中的儲存格送入 cellinexcel 的陣列中
    for column in datasheet.columns:
        for cell in column:
            cellinexcel.append(cell.value)

    columnWidthArr = []
    columnWidth = []

    # 用 if 與 for 判斷 每個儲存格數值陣列 內的值，若不為 None，則將該值設為該值的長度，若為 None，則將該值設為零
    for i in range(0, len(cellinexcel)):
        if cellinexcel[i] != None:
            cellinexcel[i] = len(cellinexcel[i])
        else:
            cellinexcel[i] = 0

    # 依照 總欄數 和 每欄個數 把陣列切開
    columnWidthArr = np.reshape(cellinexcel, (rowNum+1, columnNum))

    # 用迴圈把 columnWidthArr 中各陣列裡最大的數排序出來放入 columnWidth
    for i in range(0, len(columnWidthArr)):
        columnWidthArr[i] = max(columnWidthArr[i])
        columnWidth.append(columnWidthArr[i][0])

    # 用迴圈送入各欄的寬度
    for i in range(0, len(columnWidthArr)):
        datasheet.column_dimensions[chr(65 + i)].width = columnWidth[i]+2

# ---------done-------------

wb.save(filename=dest_filename)
